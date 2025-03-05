import glob
import string

import pandas as pd
import openpyxl

# 設定「reports.xlsx」路徑
report_path = '../output/reports.xlsx'

# 取得所有非營利問卷檔案的路徑清單
filepaths = glob.glob('../data/非營利/*.xlsx')

# 開啟「112-1-調查結果總表.xlsx」活頁簿
wb = openpyxl.load_workbook(report_path)
# 指定「112-1-調查結果總表.xlsx」活頁簿所使用的Sheet
ws = wb['非營利']

# 定義各欄位於reports.xlsx所在column索引值
return_amounts = 8  # 回收份數
total_amounts = 9   # 發放份數
return_rate = 10    # 回收率
q1 = 11             # Q1
t1 = 33             # T1

# 依序載入filepath所指明的檔案
for filepath in filepaths:
    # 從檔案路徑解析出幼兒園代碼
    serial_no = filepath.split('/')[3].split('_')[0]
    # 找出該serial_no在ws中的位置
    target_row = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == serial_no:
                target_row = cell.row
    # 讀取問卷excel檔
    qdata = pd.read_excel(filepath)
    # 轉換為dataframe
    df = pd.DataFrame(qdata)
    # 取得資料筆數（回收問卷數）
    amount = df.shape[0]
    # 求算各題平均得分
    avg_scores = df[df.columns[7:27]].mean()
    # 寫入回收份數
    ws.cell(target_row, return_amounts).value = amount
    # 求算回收率
    rate = round(amount / ws.cell(target_row, total_amounts).value, 4)
    # 寫入回收率
    ws.cell(target_row, return_rate).value = rate
    # 寫入各題的平均分數
    for i in range(q1, q1+20):
        ws.cell(target_row, i).value = round(avg_scores.iloc[i-11], 2)
    # 文字回應內容
    opinion = ""
    for i in range(27, 31):
        # 取得文字題回應（去除重複項目）
        feedbacks = df[df.columns[i]].unique()
        for feedback in feedbacks:
            if pd.isna(feedback) or pd.isnull(feedback) or feedback == "無" or feedback == "無意見" or feedback == "沒有" or feedback == "沒有意見":
                continue
            if opinion == "":
                opinion = feedback
            else:
                opinion = opinion + '\r' + feedback
        if opinion == "":
            opinion = "無"
        ws.cell(target_row, i+6).value = opinion
        opinion = ""
        del feedbacks


# 將結果存入「reports.xlsx」
wb.save(report_path)

# 關閉「reports.xlsx」
wb.close()
